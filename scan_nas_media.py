"""Durchsucht Bilder/Videos auf dem NAS und listet je Verzeichnis die Anzahl
auf (analog zu scan_phone_media.py, aber fuer das Handy via MTP).

NAS-Pfade sind normale Dateisystempfade, daher reicht hier pathlib direkt -
keine Shell-COM-Schnittstelle wie beim per MTP verbundenen Handy noetig.

Durchsucht wird \\DS923Plus\PhoneMirror\<name>\fotos\DCIM, rekursiv.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from scan_phone_media import classify

DEFAULT_NAME = "stefan"
NAS_BASE = r"\\DS923Plus\PhoneMirror"

REPO_ROOT = Path(__file__).resolve().parent
RESULTS_DIR = REPO_ROOT / "results"

logger = logging.getLogger("scan_nas_media")


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-8s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        stream=sys.stdout,
    )


def prompt_choice(label: str, default: str) -> str:
    answer = input(f"{label}? [{default}]: ").strip()
    return answer or default


def scan_folder(folder: Path, relative_path: str, results: list[dict], skip_hidden: bool) -> None:
    try:
        entries = sorted(folder.iterdir(), key=lambda p: p.name)
    except OSError as exc:
        logger.warning("Verzeichnis konnte nicht gelesen werden: %s (%s)", relative_path, exc)
        return

    images = 0
    videos = 0
    subfolders = []

    for entry in entries:
        if skip_hidden and entry.name.startswith("."):
            continue
        if entry.is_dir():
            subfolders.append(entry)
            continue
        kind = classify(entry.name)
        if kind == "image":
            images += 1
        elif kind == "video":
            videos += 1

    if images or videos:
        results.append({"path": relative_path, "images": images, "videos": videos})
        logger.info("%-70s Bilder: %5d  Videos: %5d", relative_path, images, videos)

    for entry in subfolders:
        sub_path = f"{relative_path}/{entry.name}"
        scan_folder(entry, sub_path, results, skip_hidden)


def write_excel(path: Path, results: list[dict]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Verzeichnisse"

    headers = ["Verzeichnis", "Bilder", "Videos"]
    sheet.append(headers)
    for result in results:
        sheet.append([result["path"], result["images"], result["videos"]])

    widths = [len(headers[0]), len(headers[1]), len(headers[2])]
    for result in results:
        widths[0] = max(widths[0], len(result["path"]))
    for col_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = width + 2

    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Durchsucht Bilder/Videos auf dem NAS pro Verzeichnis und schreibt ein Excel."
    )
    parser.add_argument("--name", default=None, help="Person, deren NAS-Ordner durchsucht wird (Default: stefan)")
    parser.add_argument(
        "--include-hidden",
        action="store_true",
        help="Auch versteckte Ordner durchsuchen (z.B. Syncthing-interne .stfolder/.stversions)",
    )
    args = parser.parse_args()

    setup_logging()

    name = args.name or prompt_choice("Name", DEFAULT_NAME)
    root_path = Path(NAS_BASE) / name / "fotos" / "DCIM"

    logger.info("Name: %s", name)
    logger.info("Durchsuche: %s", root_path)

    if not root_path.exists():
        logger.error("Pfad nicht erreichbar: %s", root_path)
        return 1

    results: list[dict] = []
    scan_folder(root_path, "DCIM", results, skip_hidden=not args.include_hidden)

    total_images = sum(r["images"] for r in results)
    total_videos = sum(r["videos"] for r in results)
    logger.info(
        "Fertig. %d Verzeichnisse mit Medien, insgesamt %d Bilder, %d Videos.",
        len(results), total_images, total_videos,
    )

    RESULTS_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = RESULTS_DIR / f"nas_{name}_{timestamp}.json"
    output_file.write_text(
        json.dumps(
            {
                "name": name,
                "root_path": str(root_path),
                "scanned_at": timestamp,
                "total_images": total_images,
                "total_videos": total_videos,
                "directories": results,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    logger.info("Ergebnis gespeichert: %s", output_file)

    excel_file = RESULTS_DIR / f"nas_{name}_{timestamp}.xlsx"
    write_excel(excel_file, results)
    logger.info("Excel gespeichert: %s", excel_file)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
