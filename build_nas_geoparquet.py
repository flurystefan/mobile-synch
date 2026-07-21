"""Erzeugt pro Verzeichnis auf dem NAS ein GeoParquet mit allen Bildern, die
GPS-Koordinaten im EXIF haben.

Jede Zeile enthaelt: Punkt-Geometrie (aus GPS, EPSG:4326), EXIF-Metadaten
(Aufnahmedatum, Kamera, Objektiv, Belichtung, ...), ein eingebettetes
JPEG-Thumbnail (lange Seite max. 600px, EXIF-Rotation beruecksichtigt) und
den vollen Pfad zum Original.

Kann in ArcGIS Pro (ab 3.2) direkt als Punkt-Feature-Layer geladen werden.
Hinweis: Das Thumbnail liegt als rohes Blob-Feld in der Tabelle. Damit es in
einem ArcGIS Pro Pop-up automatisch als Bild erscheint, muss die Blob-Spalte
i.d.R. erst in ein Raster-Feld einer Geodatabase-Tabelle konvertiert werden
(z.B. per arcpy) - reines Laden des Parquet zeigt das Bild nicht automatisch
im Popup an.

Bilder ohne GPS-Koordinaten werden nicht ins GeoParquet aufgenommen, sondern
in einer Excel-Datei aufgelistet.

Durchsucht wird das Verzeichnis unter json/<person>-mobile-dirs.json ->
["nas-speicher"][<nas-key>] (Default nas-key: DCIM), rekursiv. Pro Verzeichnis
(nur dessen direkte Bilder, nicht die der Unterordner) entsteht eine eigene
.parquet-Datei, deren Pfad die Quellordner-Struktur spiegelt.

Zusaetzlich entsteht pro Verzeichnis (mit demselben Pfad-Spiegel wie beim
GeoParquet):
- eine File Geodatabase mit einer Punkt-Feature-Class ("photos"), bei der
  die Thumbnails als Attachments hinterlegt sind - damit zeigt ArcGIS Pro
  das Bild automatisch im Popup an (bei reinem GeoParquet mit Blob-Feld
  geht das nicht).
- ein ArcGIS Pro Projekt mit einer Map, in der die Feature Class aus der
  zugehoerigen FGDB bereits geladen ist.

Alle drei (geoparquet/fgdb/AGP) werden direkt unter
<output-nas-base>/<person>/gis/{geoparquet,fgdb,AGP} geschrieben (Default
fuer output-nas-base: \\\\DS923Plus\\PhoneMirror) - nicht lokal, damit z.B.
die Fotokarten-Webapp (Docker auf dem NAS) direkt darauf zugreifen kann.
Nur der "ohne GPS/EXIF"-Report (JSON+Excel) landet lokal im results/-Ordner.

Benoetigt arcpy, also eine Python-Umgebung mit ArcGIS Pro Lizenz.
"""

from __future__ import annotations

import argparse
import datetime
import json
import logging
import shutil
import sys
from io import BytesIO
from pathlib import Path

import arcpy
import geopandas as gpd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps
from shapely.geometry import Point

from exif_utils import extract_camera_metadata, extract_datetime_original, extract_gps
from scan_phone_media import classify

DEFAULT_PERSON = "stefan"
DEFAULT_NAS_KEY = "DCIM"
DEFAULT_OUTPUT_NAS_BASE = r"\\DS923Plus\PhoneMirror"
THUMBNAIL_MAX_SIZE = 600

GDB_FEATURE_CLASS = "photos"
GDB_FIELDS = [
    ("filename", "TEXT", 255),
    ("directory", "TEXT", 500),
    ("full_path", "TEXT", 500),
    ("latitude", "DOUBLE", None),
    ("longitude", "DOUBLE", None),
    ("altitude_m", "DOUBLE", None),
    ("datetime_original", "TEXT", 30),
    ("width", "LONG", None),
    ("height", "LONG", None),
    ("camera_make", "TEXT", 100),
    ("camera_model", "TEXT", 100),
    ("orientation", "LONG", None),
    ("lens_model", "TEXT", 150),
    ("focal_length_mm", "DOUBLE", None),
    ("f_number", "DOUBLE", None),
    ("exposure_time_s", "DOUBLE", None),
    ("iso_speed", "LONG", None),
]

REPO_ROOT = Path(__file__).resolve().parent
JSON_DIR = REPO_ROOT / "json"
RESULTS_DIR = REPO_ROOT / "results"

logger = logging.getLogger("build_nas_geoparquet")


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-8s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        stream=sys.stdout,
    )


def available_persons() -> list[str]:
    return sorted(
        p.name[: -len("-mobile-dirs.json")]
        for p in JSON_DIR.glob("*-mobile-dirs.json")
    )


def prompt_choice(label: str, default: str, choices: list[str] | None = None) -> str:
    hint = f" (verfuegbar: {', '.join(choices)})" if choices else ""
    answer = input(f"{label}? [{default}]{hint}: ").strip()
    return answer or default


def load_nas_root(person: str, nas_key: str) -> Path:
    dirs_file = JSON_DIR / f"{person}-mobile-dirs.json"
    if not dirs_file.exists():
        raise RuntimeError(f"Datei nicht gefunden: {dirs_file}")

    data = json.loads(dirs_file.read_text(encoding="utf-8"))
    nas_speicher = data.get("nas-speicher", {})
    if nas_key not in nas_speicher:
        raise RuntimeError(
            f"Schluessel '{nas_key}' nicht unter 'nas-speicher' in {dirs_file} gefunden. "
            f"Verfuegbar: {', '.join(nas_speicher) or 'keine'}"
        )
    return Path(nas_speicher[nas_key])


def mirrored_path(root: Path, directory: Path, nas_root: Path, suffix: str) -> Path:
    """Spiegelt einen NAS-Verzeichnispfad unter root, mit gegebener Datei-Endung."""
    relative = directory.relative_to(nas_root)
    if str(relative) == ".":
        return root / f"{directory.name}{suffix}"
    return root / relative.with_suffix(suffix)


def mirrored_dir(root: Path, directory: Path, nas_root: Path) -> Path:
    """Wie mirrored_path, aber fuer einen reinen Ordnernamen (kein Datei-Suffix)."""
    relative = directory.relative_to(nas_root)
    if str(relative) == ".":
        return root / directory.name
    return root / relative


def write_gdb(gdb_dir: Path, rows: list[dict]) -> str:
    """Erzeugt (bzw. ersetzt) eine File Geodatabase mit einer Punkt-Feature-Class
    (inkl. Attachments fuer die Thumbnails) fuer genau ein Verzeichnis.

    Feature Class und Attachment-Tabelle haengen ueber eine Relationship-Class
    zusammen, daher muessen Aenderungen in einer Edit-Session laufen.
    """
    if arcpy.Exists(str(gdb_dir)):
        arcpy.management.Delete(str(gdb_dir))

    gdb_dir.parent.mkdir(parents=True, exist_ok=True)
    arcpy.management.CreateFileGDB(str(gdb_dir.parent), gdb_dir.name)

    fc_path = str(gdb_dir / GDB_FEATURE_CLASS)
    arcpy.management.CreateFeatureclass(
        str(gdb_dir), GDB_FEATURE_CLASS, "POINT",
        spatial_reference=arcpy.SpatialReference(4326),
    )
    field_names = [name for name, _, _ in GDB_FIELDS]
    for name, field_type, length in GDB_FIELDS:
        if length:
            arcpy.management.AddField(fc_path, name, field_type, field_length=length)
        else:
            arcpy.management.AddField(fc_path, name, field_type)
    arcpy.management.EnableAttachments(fc_path)

    editor = arcpy.da.Editor(str(gdb_dir))
    editor.startEditing(False, False)
    editor.startOperation()
    try:
        oids = []
        with arcpy.da.InsertCursor(fc_path, ["SHAPE@XY"] + field_names) as cursor:
            for row in rows:
                values = [(row["longitude"], row["latitude"])] + [row.get(name) for name in field_names]
                oids.append(cursor.insertRow(values))

        attach_table = f"{fc_path}__ATTACH"
        with arcpy.da.InsertCursor(
            attach_table, ["REL_OBJECTID", "CONTENT_TYPE", "ATT_NAME", "DATA_SIZE", "DATA"]
        ) as attach_cursor:
            for oid, row in zip(oids, rows):
                thumbnail = row["thumbnail"]
                attach_cursor.insertRow(
                    (oid, "image/jpeg", row["filename"], len(thumbnail), memoryview(thumbnail))
                )
    except Exception:
        editor.stopOperation()
        editor.stopEditing(False)
        raise
    else:
        editor.stopOperation()
        editor.stopEditing(True)

    return fc_path


def relocate_import_log(gdb_dir: Path) -> None:
    """arcpy.mp.Map.createMap() schreibt beim Laden der Basemap ein 'ImportLog'
    neben die als default_database angegebene FGDB. Das wollen wir nicht neben
    den FGDBs haben, sondern gesammelt im obersten results-Verzeichnis."""
    import_log = gdb_dir.parent / "ImportLog"
    if not import_log.exists():
        return

    target = RESULTS_DIR / "ImportLog"
    target.mkdir(exist_ok=True)
    for item in import_log.iterdir():
        shutil.move(str(item), str(target / item.name))
    import_log.rmdir()


def create_project(project_dir: Path, fc_path: str) -> None:
    """Erzeugt (bzw. ersetzt) ein ArcGIS Pro Projekt in einem eigenen
    Unterordner (project_dir/<name>.aprx), mit einer Map, in der die Feature
    Class aus der zugehoerigen FGDB bereits geladen ist.

    Ein eigener Ordner pro Projekt ist noetig, weil ArcGIS Pro rund um die
    .aprx weitere Dateien ablegt (Toolbox, Backups, Index) - bei vielen
    Projekten in einem einzigen Ordner wuerde das schnell unuebersichtlich.
    """
    if project_dir.exists():
        shutil.rmtree(project_dir)

    project_dir.parent.mkdir(parents=True, exist_ok=True)
    project_name = project_dir.name
    gdb_dir = Path(fc_path).parent
    arcpy.mp.CreateArcGISProject(
        str(project_dir.parent), project_name,
        create_parent_folder=True,
        default_database=str(gdb_dir),
    )

    aprx_path = project_dir / f"{project_name}.aprx"
    aprx = arcpy.mp.ArcGISProject(str(aprx_path))
    map_obj = aprx.createMap("Fotos")
    map_obj.addDataFromPath(fc_path)
    aprx.save()

    relocate_import_log(gdb_dir)


def process_image(entry: Path) -> tuple[dict | None, str | None]:
    """Liest EXIF+GPS und erzeugt ein Thumbnail in einem einzigen Datei-Open.

    Gibt (row, None) bei Erfolg zurueck, (None, grund) wenn das Bild nicht
    aufgenommen werden kann (kein GPS oder Lesefehler).
    """
    try:
        with Image.open(entry) as img:
            exif = img.getexif()
            gps = extract_gps(exif)
            if gps is None:
                return None, "no_gps"

            width, height = img.size
            dt = extract_datetime_original(exif)
            metadata = extract_camera_metadata(exif)

            oriented = ImageOps.exif_transpose(img)
            oriented.thumbnail((THUMBNAIL_MAX_SIZE, THUMBNAIL_MAX_SIZE), Image.LANCZOS)
            buffer = BytesIO()
            oriented.convert("RGB").save(buffer, format="JPEG", quality=85)
            thumbnail = buffer.getvalue()
    except Exception as exc:
        logger.warning("Konnte Bild nicht verarbeiten: %s (%s)", entry, exc)
        return None, "error"

    latitude, longitude, altitude = gps
    row = {
        "geometry": Point(longitude, latitude),
        "filename": entry.name,
        "directory": str(entry.parent),
        "full_path": str(entry),
        "latitude": latitude,
        "longitude": longitude,
        "altitude_m": altitude,
        "datetime_original": dt.isoformat() if dt else None,
        "width": width,
        "height": height,
        "thumbnail": thumbnail,
        **metadata,
    }
    return row, None


def build_rows_for_directory(directory: Path, entries: list[Path], no_gps: list[dict]) -> list[dict]:
    rows = []
    for entry in entries:
        if entry.is_dir() or entry.name.startswith("."):
            continue
        if classify(entry.name) != "image":
            continue

        row, _ = process_image(entry)
        if row is None:
            no_gps.append({"path": str(directory), "name": entry.name})
            continue
        rows.append(row)
    return rows


def process_directory(
    directory: Path,
    nas_root: Path,
    output_root: Path,
    fgdb_root: Path,
    aprx_root: Path,
    stats: dict,
    no_gps: list[dict],
    skip_existing: bool,
) -> None:
    try:
        entries = sorted(directory.iterdir(), key=lambda p: p.name)
    except OSError as exc:
        logger.warning("Verzeichnis konnte nicht gelesen werden: %s (%s)", directory, exc)
        return

    output_path = mirrored_path(output_root, directory, nas_root, ".parquet")

    if skip_existing and output_path.exists():
        stats["skipped"] += 1
    else:
        rows = build_rows_for_directory(directory, entries, no_gps)
        if rows:
            output_path.parent.mkdir(parents=True, exist_ok=True)

            gdf = gpd.GeoDataFrame(rows, geometry="geometry", crs="EPSG:4326")
            gdf.to_parquet(output_path)

            gdb_dir = mirrored_path(fgdb_root, directory, nas_root, ".gdb")
            fc_path = write_gdb(gdb_dir, rows)

            project_dir = mirrored_dir(aprx_root, directory, nas_root)
            create_project(project_dir, fc_path)

            stats["directories"] += 1
            stats["images"] += len(rows)
            logger.info("GeoParquet geschrieben: %s (%d Bild(er))", output_path, len(rows))

    for entry in entries:
        if entry.is_dir() and not entry.name.startswith("."):
            process_directory(entry, nas_root, output_root, fgdb_root, aprx_root, stats, no_gps, skip_existing)


def write_no_gps_excel(path: Path, no_gps: list[dict]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ohne GPS"

    headers = ["Verzeichnis", "Dateiname"]
    sheet.append(headers)
    for row in no_gps:
        full_path = Path(row["path"]) / row["name"]
        sheet.append([row["path"], str(full_path)])

        dir_cell = sheet.cell(row=sheet.max_row, column=1)
        dir_cell.hyperlink = Path(row["path"]).as_uri()
        dir_cell.font = Font(color="0000FF", underline="single")

        name_cell = sheet.cell(row=sheet.max_row, column=2)
        name_cell.hyperlink = full_path.as_uri()
        name_cell.font = Font(color="0000FF", underline="single")

    widths = [len(headers[0]), len(headers[1])]
    for row in no_gps:
        widths[0] = max(widths[0], len(row["path"]))
        widths[1] = max(widths[1], len(str(Path(row["path"]) / row["name"])))
    for col_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = width + 2

    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Erzeugt pro Verzeichnis ein GeoParquet mit georeferenzierten Bildern (aus EXIF-GPS)."
    )
    parser.add_argument("--person", default=None, help="Person, deren Verzeichnisliste genutzt wird (z.B. stefan, pia)")
    parser.add_argument("--nas-key", default=DEFAULT_NAS_KEY, help="Schluessel unter 'nas-speicher' im json (Default: DCIM)")
    parser.add_argument(
        "--output-nas-base",
        default=DEFAULT_OUTPUT_NAS_BASE,
        help=f"NAS-Basispfad, unter dem <person>/gis/{{geoparquet,fgdb,AGP}} geschrieben wird (Default: {DEFAULT_OUTPUT_NAS_BASE})",
    )
    parser.add_argument(
        "--dirs",
        default=None,
        help="Kommagetrennte Liste von Verzeichnissen (relativ zum NAS-Root), die statt des "
        "gesamten Baums verarbeitet werden. Bereits vorhandene GeoParquets werden dabei ueberschrieben.",
    )
    parser.add_argument(
        "--only-missing",
        action=argparse.BooleanOptionalAction,
        default=None,
        help="Nur fehlende GeoParquets anlegen, vorhandene ueberspringen (Default: ja). "
        "Mit --no-only-missing werden auch vorhandene neu erstellt. Ohne Angabe wird gefragt.",
    )
    args = parser.parse_args()

    setup_logging()

    person = args.person or prompt_choice("Person", DEFAULT_PERSON, available_persons())

    if args.only_missing is None:
        answer = prompt_choice("Nur fehlende GeoParquets anlegen (only-missing)", "ja")
        only_missing = answer.strip().lower() not in ("nein", "n", "no", "false", "0")
    else:
        only_missing = args.only_missing

    if args.dirs is None:
        answer = prompt_choice(
            "Nur bestimmte Verzeichnisse bearbeiten (kommagetrennt, relativ zum NAS-Root)",
            "alle",
        )
        args.dirs = None if answer.strip().lower() in ("", "alle", "all") else answer

    logger.info(
        "Person: %s  NAS-Key: %s  Only-Missing: %s  Dirs: %s  Ausgabe-Basis: %s",
        person, args.nas_key, only_missing, args.dirs or "alle", args.output_nas_base,
    )

    try:
        nas_root = load_nas_root(person, args.nas_key)
    except RuntimeError as exc:
        logger.error(str(exc))
        return 1

    if not nas_root.exists():
        logger.error("NAS-Pfad nicht erreichbar: %s", nas_root)
        return 1

    RESULTS_DIR.mkdir(exist_ok=True)
    person_root = Path(args.output_nas_base) / person / "gis"
    output_root = person_root / "geoparquet"
    fgdb_root = person_root / "fgdb"
    aprx_root = person_root / "AGP"

    stats = {"directories": 0, "images": 0, "skipped": 0}
    no_gps: list[dict] = []

    if args.dirs:
        for raw in args.dirs.split(","):
            relative = raw.strip()
            if not relative:
                continue
            target = nas_root / relative
            if not target.exists():
                logger.error("Verzeichnis nicht gefunden: %s", target)
                continue
            logger.info("Verarbeite (erzwungen): %s", target)
            process_directory(
                target, nas_root, output_root, fgdb_root, aprx_root, stats, no_gps, skip_existing=False
            )
    else:
        process_directory(
            nas_root, nas_root, output_root, fgdb_root, aprx_root, stats, no_gps, skip_existing=only_missing
        )

    logger.info(
        "Fertig. %d GeoParquet-Datei(en) mit insgesamt %d Bild(ern) geschrieben, "
        "%d uebersprungen (bereits vorhanden), %d Bild(er) ohne GPS.",
        stats["directories"], stats["images"], stats["skipped"], len(no_gps),
    )

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = RESULTS_DIR / f"nas_{person}_no_gps_{timestamp}.xlsx"
    write_no_gps_excel(excel_file, no_gps)
    logger.info("Excel (ohne GPS) gespeichert: %s", excel_file)
    logger.info("GeoParquet-Dateien liegen unter: %s", output_root)
    logger.info("File Geodatabases (mit Popup-Attachments) liegen unter: %s", fgdb_root)
    logger.info("ArcGIS Pro Projekte liegen unter: %s", aprx_root)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
