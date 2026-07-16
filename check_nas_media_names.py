"""Prueft und normalisiert Bild-/Video-Dateinamen auf dem NAS gegen das Muster
YYYYMMDD_HHMMSS[.ext], mit " (N)" Suffix bei Kollisionen.

Anders als beim Handy (MTP) sind NAS-Pfade normale Dateisystempfade, daher
wird hier direkt mit pathlib gearbeitet. Der Zielname wird aus den EXIF-Daten
(Aufnahmedatum) der Datei abgeleitet, nicht aus dem Datei-Zeitstempel des
Dateisystems. Dateien ohne lesbares EXIF-Datum (z.B. Videos, Screenshots)
werden NICHT automatisch umbenannt, sondern in einer Excel-Datei aufgelistet.

Durchsucht wird das Verzeichnis unter json/<person>-mobile-dirs.json ->
["nas-speicher"][<nas-key>] (Default nas-key: DCIM), rekursiv.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from exif_utils import get_exif_datetime
from media_naming import is_valid_name
from scan_phone_media import classify

DEFAULT_PERSON = "stefan"
DEFAULT_NAS_KEY = "DCIM"
DEFAULT_MODE = "dry_run"

REPO_ROOT = Path(__file__).resolve().parent
JSON_DIR = REPO_ROOT / "json"
RESULTS_DIR = REPO_ROOT / "results"

logger = logging.getLogger("check_nas_media_names")


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-8s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        stream=sys.stdout,
    )


def available_persons() -> list[str]:
    return sorted(
        p.name[: -len("-mobile-dirs.json")]
        for p in JSON_DIR.glob("*-mobile-dirs.json")
    )


def prompt_choice(label: str, default: str, choices: list[str] | None = None) -> str:
    hint = f" (verfuegbar: {', '.join(choices)})" if choices else ""
    answer = input(f"{label}? [{default}]{hint}: ").strip()
    return answer or default


def load_nas_root(person: str, nas_key: str) -> Path:
    dirs_file = JSON_DIR / f"{person}-mobile-dirs.json"
    if not dirs_file.exists():
        raise RuntimeError(f"Datei nicht gefunden: {dirs_file}")

    data = json.loads(dirs_file.read_text(encoding="utf-8"))
    nas_speicher = data.get("nas-speicher", {})
    if nas_key not in nas_speicher:
        raise RuntimeError(
            f"Schluessel '{nas_key}' nicht unter 'nas-speicher' in {dirs_file} gefunden. "
            f"Verfuegbar: {', '.join(nas_speicher) or 'keine'}"
        )
    return Path(nas_speicher[nas_key])


def next_free_name(directory: Path, stem: str, suffix: str, used_names: set[str]) -> str:
    candidate = f"{stem}{suffix}"
    if candidate not in used_names and not (directory / candidate).exists():
        return candidate

    n = 1
    while True:
        candidate = f"{stem} ({n}){suffix}"
        if candidate not in used_names and not (directory / candidate).exists():
            return candidate
        n += 1


def process_directory(directory: Path, mode: str, skip_hidden: bool, results: dict) -> None:
    try:
        entries = sorted(directory.iterdir(), key=lambda p: p.name)
    except OSError as exc:
        logger.warning("Verzeichnis konnte nicht gelesen werden: %s (%s)", directory, exc)
        return

    used_names = {p.name for p in entries}

    for entry in entries:
        if skip_hidden and entry.name.startswith("."):
            continue

        if entry.is_dir():
            process_directory(entry, mode, skip_hidden, results)
            continue

        kind = classify(entry.name)
        if kind is None:
            continue

        results["checked"] += 1
        if is_valid_name(entry.stem):
            continue

        exif_dt = get_exif_datetime(entry) if kind == "image" else None
        if exif_dt is None:
            results["no_exif"].append(
                {"path": str(entry.parent), "name": entry.name, "type": kind}
            )
            logger.info("Kein EXIF-Datum: %s", entry)
            continue

        new_stem = exif_dt.strftime("%Y%m%d_%H%M%S")
        new_name = next_free_name(directory, new_stem, entry.suffix, used_names)
        used_names.add(new_name)

        results["renamed"].append(
            {"path": str(entry.parent), "old_name": entry.name, "new_name": new_name}
        )

        if mode == "fix_names":
            entry.rename(entry.with_name(new_name))
            logger.info("Umbenannt: %s -> %s", entry.name, new_name)
        else:
            logger.info("[dry_run] Wuerde umbenennen: %s -> %s", entry, new_name)


def write_no_exif_excel(path: Path, no_exif: list[dict]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ohne EXIF"

    headers = ["Verzeichnis", "Dateiname", "Typ"]
    sheet.append(headers)
    for row in no_exif:
        full_path = Path(row["path"]) / row["name"]
        sheet.append([row["path"], str(full_path), row["type"]])

        dir_cell = sheet.cell(row=sheet.max_row, column=1)
        dir_cell.hyperlink = Path(row["path"]).as_uri()
        dir_cell.font = Font(color="0000FF", underline="single")

        name_cell = sheet.cell(row=sheet.max_row, column=2)
        name_cell.hyperlink = full_path.as_uri()
        name_cell.font = Font(color="0000FF", underline="single")

    widths = [len(headers[0]), len(headers[1]), len(headers[2])]
    for row in no_exif:
        widths[0] = max(widths[0], len(row["path"]))
        widths[1] = max(widths[1], len(str(Path(row["path"]) / row["name"])))
    for col_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = width + 2

    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Prueft/normalisiert Bild-/Video-Dateinamen auf dem NAS gegen YYYYMMDD_HHMMSS."
    )
    parser.add_argument("--person", default=None, help="Person, deren Verzeichnisliste genutzt wird (z.B. stefan, pia)")
    parser.add_argument("--mode", choices=["dry_run", "fix_names"], default=None, help="dry_run (nur anzeigen) oder fix_names (umbenennen)")
    parser.add_argument("--nas-key", default=DEFAULT_NAS_KEY, help="Schluessel unter 'nas-speicher' im json (Default: DCIM)")
    args = parser.parse_args()

    setup_logging()

    person = args.person or prompt_choice("Person", DEFAULT_PERSON, available_persons())
    mode = args.mode or prompt_choice("Modus (dry_run/fix_names)", DEFAULT_MODE)
    logger.info("Person: %s  Modus: %s  NAS-Key: %s", person, mode, args.nas_key)

    try:
        nas_root = load_nas_root(person, args.nas_key)
    except RuntimeError as exc:
        logger.error(str(exc))
        return 1

    if not nas_root.exists():
        logger.error("NAS-Pfad nicht erreichbar: %s", nas_root)
        return 1

    results: dict = {"checked": 0, "renamed": [], "no_exif": []}
    process_directory(nas_root, mode, skip_hidden=True, results=results)

    logger.info(
        "Fertig. %d Datei(en) geprueft, %d %s, %d ohne EXIF-Datum.",
        results["checked"],
        len(results["renamed"]),
        "umbenannt" if mode == "fix_names" else "wuerden umbenannt (dry_run)",
        len(results["no_exif"]),
    )

    RESULTS_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_file = RESULTS_DIR / f"nas_{person}_name_check_{timestamp}.json"
    json_file.write_text(
        json.dumps(
            {
                "person": person,
                "mode": mode,
                "nas_root": str(nas_root),
                "checked_at": timestamp,
                "checked": results["checked"],
                "renamed": results["renamed"],
                "no_exif": results["no_exif"],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    logger.info("Ergebnis gespeichert: %s", json_file)

    excel_file = RESULTS_DIR / f"nas_{person}_no_exif_{timestamp}.xlsx"
    write_no_exif_excel(excel_file, results["no_exif"])
    logger.info("Excel (ohne EXIF) gespeichert: %s", excel_file)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
