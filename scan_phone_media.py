"""Durchsucht ein per MTP verbundenes Android-Handy (sichtbar unter "Dieser PC")
nach Bild- und Videodateien und listet je Verzeichnis die Anzahl auf.

MTP-Geraete haben keinen Laufwerksbuchstaben, daher funktionieren normale
Dateizugriffe (os.walk, open, ...) hier nicht. Der Zugriff erfolgt ueber die
Windows Shell-COM-Schnittstelle (win32com), genau wie der Explorer sie nutzt.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import pywintypes
import win32com.client
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp",
    ".heic", ".heif", ".dng", ".raw", ".tif", ".tiff",
}
VIDEO_EXTENSIONS = {
    ".mp4", ".mov", ".avi", ".mkv", ".3gp", ".m4v", ".wmv", ".webm",
}

DEFAULT_DEVICE_NAME = "S23 Ultra von Stefan"
DEFAULT_STORAGE_NAME = "Interner Speicher"
CSIDL_DRIVES = 17  # "Dieser PC"

REPO_ROOT = Path(__file__).resolve().parent
RESULTS_DIR = REPO_ROOT / "results"

logger = logging.getLogger("scan_phone_media")


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-8s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        stream=sys.stdout,
    )


def get_root_folder(device_name: str, storage_name: str):
    shell = win32com.client.Dispatch("Shell.Application")
    this_pc = shell.Namespace(CSIDL_DRIVES)
    if this_pc is None:
        raise RuntimeError("Konnte 'Dieser PC' nicht oeffnen.")

    device_item = next((item for item in this_pc.Items() if item.Name == device_name), None)
    if device_item is None:
        available = ", ".join(item.Name for item in this_pc.Items())
        raise RuntimeError(
            f"Geraet '{device_name}' nicht unter 'Dieser PC' gefunden. "
            f"Ist das Handy eingesteckt und entsperrt? Verfuegbar: {available}"
        )

    storage_item = next(
        (item for item in device_item.GetFolder.Items() if item.Name == storage_name), None
    )
    if storage_item is None:
        available = ", ".join(item.Name for item in device_item.GetFolder.Items())
        raise RuntimeError(
            f"Ordner '{storage_name}' nicht in '{device_name}' gefunden. Verfuegbar: {available}"
        )

    return storage_item.GetFolder


def classify(name: str) -> str | None:
    suffix = Path(name).suffix.lower()
    if suffix in IMAGE_EXTENSIONS:
        return "image"
    if suffix in VIDEO_EXTENSIONS:
        return "video"
    return None


def scan_folder(folder, relative_path: str, results: list[dict], skip_hidden: bool) -> None:
    images = 0
    videos = 0
    subfolders = []

    for item in folder.Items():
        if skip_hidden and item.Name.startswith("."):
            logger.debug("Uebersprungen (versteckt): %s/%s", relative_path, item.Name)
            continue
        if item.IsFolder:
            subfolders.append(item)
            continue
        kind = classify(item.Name)
        if kind == "image":
            images += 1
        elif kind == "video":
            videos += 1

    if images or videos:
        results.append({"path": relative_path, "images": images, "videos": videos})
        logger.info("%-70s Bilder: %5d  Videos: %5d", relative_path, images, videos)

    for item in subfolders:
        sub_path = f"{relative_path}/{item.Name}"
        try:
            scan_folder(item.GetFolder, sub_path, results, skip_hidden)
        except pywintypes.com_error as exc:
            logger.warning("Verzeichnis konnte nicht gelesen werden: %s (%s)", sub_path, exc)


def write_excel(path: Path, results: list[dict]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Verzeichnisse"

    headers = ["Verzeichnis", "Bilder", "Videos"]
    sheet.append(headers)
    for result in results:
        sheet.append([result["path"], result["images"], result["videos"]])

    widths = [len(headers[0])] + [0, 0]
    for result in results:
        widths[0] = max(widths[0], len(result["path"]))
    widths[1] = len(headers[1])
    widths[2] = len(headers[2])
    for col_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = width + 2

    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Durchsucht das Handy (MTP) nach Bildern/Videos pro Verzeichnis."
    )
    parser.add_argument("--device-name", default=DEFAULT_DEVICE_NAME, help="Name unter 'Dieser PC'")
    parser.add_argument("--storage-name", default=DEFAULT_STORAGE_NAME, help="Speicherordner auf dem Geraet")
    parser.add_argument(
        "--include-hidden",
        action="store_true",
        help="Auch versteckte Ordner durchsuchen (z.B. Syncthing-interne .stfolder/.stversions)",
    )
    args = parser.parse_args()

    setup_logging()
    logger.info("Suche Geraet '%s' unter 'Dieser PC' ...", args.device_name)

    try:
        root_folder = get_root_folder(args.device_name, args.storage_name)
    except RuntimeError as exc:
        logger.error(str(exc))
        return 1

    results: list[dict] = []
    scan_folder(root_folder, args.storage_name, results, skip_hidden=not args.include_hidden)

    total_images = sum(r["images"] for r in results)
    total_videos = sum(r["videos"] for r in results)
    logger.info(
        "Fertig. %d Verzeichnisse mit Medien, insgesamt %d Bilder, %d Videos.",
        len(results), total_images, total_videos,
    )

    RESULTS_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_device_name = args.device_name.replace(" ", "_")
    output_file = RESULTS_DIR / f"{safe_device_name}_{timestamp}.json"
    output_file.write_text(
        json.dumps(
            {
                "device_name": args.device_name,
                "storage_name": args.storage_name,
                "scanned_at": timestamp,
                "total_images": total_images,
                "total_videos": total_videos,
                "directories": results,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    logger.info("Ergebnis gespeichert: %s", output_file)

    excel_file = RESULTS_DIR / f"{safe_device_name}_{timestamp}.xlsx"
    write_excel(excel_file, results)
    logger.info("Excel gespeichert: %s", excel_file)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
